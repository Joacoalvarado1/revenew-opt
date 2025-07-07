from django.shortcuts import render
from .optimization import OptimizationModel
from .dataloader import DataLoader
from .models import OptimizationResult
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

import pandas as pd
import random


def upload_csv(request):
    data = ''
    error_message = ''

    if request.method == 'POST':
        if 'test' in request.POST:
            # Datos simulados
            df = pd.DataFrame({
                'Product_A_Production_Time_Machine_1': [round(random.uniform(1.0, 2.0), 2)],
                'Product_A_Production_Time_Machine_2': [round(random.uniform(1.5, 2.5), 2)],
                'Product_B_Production_Time_Machine_1': [round(random.uniform(0.5, 1.5), 2)],
                'Product_B_Production_Time_Machine_2': [round(random.uniform(1.0, 2.0), 2)],
                'Machine_1_Available_Hours': [10.0],
                'Machine_2_Available_Hours': [12.0],
                'Price_Product_A': [random.randint(80, 120)],
                'Price_Product_B': [random.randint(60, 100)],
            })

            capacidad_a = round(random.uniform(8, 15), 2)
            capacidad_b = round(random.uniform(10, 20), 2)

            model = OptimizationModel(df, capacidad_a, capacidad_b)
            solution, total = model.solve()

            # DEBUG - para ver el diccionario solution en consola
            print("DEBUG solution dict:", solution)

            OptimizationResult.objects.create(
                capacidad_a=capacidad_a,
                capacidad_b=capacidad_b,
                producto_a=solution.get('Producto A', 0),
                producto_b=solution.get('Producto B', 0),
                ingreso_total=total,
                fecha=timezone.now()
            )

            html = "<h4>Resultado del test:</h4><ul>"
            for producto, cantidad in solution.items():
                html += f"<li><strong>{producto}</strong>: {cantidad} unidades</li>"
            html += f"</ul><p><strong>Ingreso total óptimo:</strong> {total}</p>"

            historial = OptimizationResult.objects.all().order_by('-fecha')[:10]
            return render(request, 'upload.html', {'data': html, 'historial': historial})

        elif request.FILES.get('file') and request.POST.get('capacidad_a') and request.POST.get('capacidad_b'):
            try:
                capacidad_a = float(request.POST['capacidad_a'])
                capacidad_b = float(request.POST['capacidad_b'])

                loader = DataLoader(request.FILES['file'])
                df = loader.load_and_validate()

                model = OptimizationModel(df, capacidad_a, capacidad_b)
                solution, total = model.solve()

                # DEBUG - para ver el diccionario solution en consola
                print("DEBUG solution dict:", solution)

                OptimizationResult.objects.create(
                    capacidad_a=capacidad_a,
                    capacidad_b=capacidad_b,
                    producto_a=solution.get('Producto A', 0),
                    producto_b=solution.get('Producto B', 0),
                    ingreso_total=total,
                    fecha=timezone.now()
                )

                html = "<h4>Solución óptima:</h4><ul>"
                for producto, cantidad in solution.items():
                    html += f"<li><strong>{producto}</strong>: {cantidad} unidades</li>"
                html += f"</ul><p><strong>Ingreso total óptimo:</strong> {total}</p>"

                historial = OptimizationResult.objects.all().order_by('-fecha')[:10]
                return render(request, 'upload.html', {'data': html, 'historial': historial})

            except ValueError as ve:
                error_message = f"⚠️ {str(ve)}"
            except Exception as e:
                error_message = f"⚠️ Error inesperado: {str(e)}"
        else:
            error_message = "⚠️ Debe subir un archivo CSV y completar ambas capacidades."

    historial = OptimizationResult.objects.all().order_by('-fecha')[:10]
    return render(request, 'upload.html', {'data': data, 'error': error_message, 'historial': historial})


def descargar_historial_excel(request):
    historial = OptimizationResult.objects.all().order_by('-fecha')[:50]  # puedes cambiar el límite

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"

    headers = ['Fecha', 'Capacidad A', 'Capacidad B', 'Producto A (unid)', 'Producto B (unid)', 'Ingreso total']
    ws.append(headers)

    for obj in historial:
        ws.append([
            obj.fecha.strftime('%Y-%m-%d %H:%M'),
            obj.capacidad_a,
            obj.capacidad_b,
            obj.producto_a,
            obj.producto_b,
            obj.ingreso_total
        ])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=historial_optimizacion.xlsx'
    wb.save(response)
    return response


def test_opt_model(request):
    return render(request, 'upload.html', {
        'data': 'Función de test separada no implementada aún. Usa el botón de test en la vista principal.',
        'historial': OptimizationResult.objects.all().order_by('-fecha')[:10]
    })
